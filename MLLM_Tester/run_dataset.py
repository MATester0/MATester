import os
import logging
import argparse
import openpyxl
import textwrap
from runner import Runner
from analysis import Analysis

from management.runInfo import runTimeInfo

def validate_id_exists(value):
	"""
	check the existence of id
	"""
	try:
		id = int(value)
		if id not in [1, 2, 3, 4, 5, 6, 7]:
			raise argparse.ArgumentTypeError(f"agent id {id} is not available")
		return id
	except ValueError:
		raise argparse.ArgumentTypeError(f"agent id {id} is not an interger")
	
def acquire_sheet(agent_id: int):
	if agent_id in [1]:
		return "web"
	if agent_id in [2, 3, 4]:
		return "android app"
	if agent_id in [5]:
		return "android"
	if agent_id in [6]:
		return "minecraft"
	if agent_id in [7]:
		return None
	raise ValueError(f"{agent_id} is not available")

def arg_parse():
	"""
	parse args
	"""
	parser = argparse.ArgumentParser(
	description='Automatic tester on M-agents',
	formatter_class=argparse.RawDescriptionHelpFormatter,
	epilog=textwrap.dedent('''
	example:
	# base usage: identify agent id
	python run_dataset.py -i 1"
	''')
	)

    # add necessary parameters
	required_group = parser.add_argument_group('necessary args')
	required_group.add_argument(
	    '-i', '--agent-id',
		dest="agent_id",
	    type=validate_id_exists,
	    required=True,
	    help='Agent\'s unique id, should be available in agents folder'
	)

    # parse args
	args = parser.parse_args()	
		
	return args

def main():
	""" examples:
		run_info = rn.run(1, "find a random music", None, analyzing_existing=True)
		run_info = rn.run(2, "search 'Clock' app and open it", "com.simplemobiletools.applauncher", analyzing_existing=True)
		run_info = rn.run(3, "search 'Clock' app and open it", "com.simplemobiletools.applauncher", analyzing_existing=True)
		run_info = rn.run(4, "search 'Clock' app and open it", "com.simplemobiletools.applauncher", analyzing_existing=True)
		run_info = rn.run(5, "search 'Clock' app and open it, set a 2-minute timer", None, analyzing_existing=True)
		run_info = rn.run(6, "obtain_firebrands", None, analyzing_existing=True)
		run_info = rn.run(7, None, None, analyzing_existing=True)
	"""
	args = arg_parse()
	sheet_name = acquire_sheet(args.agent_id)
	rn = Runner()
	an = Analysis()

	if sheet_name == None:
		run_info = rn.run(args.agent_id, None, None, analyzing_existing=True)
		an.run(run_info, use_existing=True)
		run_infos = [run_info]
		for id in range(1, 10): # run 10 times, test usability
			run_info_id = rn.run(args.agent_id, None, None, running_times=id, analyzing_existing=True)
			an.run(run_info_id, use_existing=True)
			run_infos.append(run_info_id)
		# analyze each pair of them to test consistency
		for id in range(9):
			run_info = run_infos[id]
			for id2 in range(id+1, 10):
				run_info_2 = run_infos[id2]
				an.run_multiple(run_info, run_info_2, use_existing=True)
	else:
		book = openpyxl.load_workbook(os.path.join(os.path.abspath("../dataset/"), "datasets.xlsx"))
		sheet = book[sheet_name]
		for r in range(2, sheet.max_row + 1):
			apk_name = sheet.cell(row = r, column=1).value
			if apk_name == "NA":
				apk_name = None
			for c in range(2, sheet.max_column + 1):
				task_name = sheet.cell(row = r, column=c).value
				if not task_name:
					continue
				# run once, test usability
				run_info = rn.run(args.agent_id, task_name, apk_name, analyzing_existing=True)
				an.run(run_info, use_existing=True)

				# run the second time, test usability and consistency
				run_info_2 = rn.run(args.agent_id, task_name, apk_name, running_times=1, analyzing_existing=True)
				an.run(run_info_2, use_existing=True)
				an.run_multiple(run_info, run_info_2, use_existing=True)
		book.close()

if __name__ == "__main__":
	# use this to run 1, 2, 3, 4, 5, 6, 7
	logging.basicConfig(level=logging.INFO)
	main()